import math

"""print('Anant Kishore')
print('o----')
print(' ||||')
print('*' * 10)

price = 10
rating = 4.9
name = 'Anant'
is_done = True

print(price,','*10)
print(rating)


full_name = "John Johny"
age = 20
is_new = True


name = input('Enter your name? ')
color = input('Enter your fav color? ')
print(name + ' likes ' + color)

birth_year = int(input('Birth year: '))
print(type(birth_year))
age = 2019 - birth_year
print(age)



wt_kgs = float(input('Weight (kgs): '))
wt_lbs = wt_kgs * 2.2570


print(wt_lbs)

course = 'Python for "Beginners"'


course = '''

Hi ANnat

how are you doing 

thanks
Anant'''

print(course)



course = 'Python for Beginners'
print(course[-1:])

course_2 = course[:]

print(course_2)



name = 'Jennifer'
print(name[1:-1])



first = 'Anant'
last = 'Kishore'

print(first +'[' + last + ']' + 'is a coder')

msg = f'{first} [{last}] is a coder'
print(msg)

course = 'Python for Beginners'
print(len(course))
print(course.upper())
print(course.lower())
print(course)

print(course.find('o'))
print(course.find('beg'))
print(course.find('Beg'))

print(course.replace('Beginners','immature'))

print('Python' in course) # returns boolean value



print( 10/3 )
print( 10//3 )
print( 10 % 3 )
print( 10 ** 3 )

x = 10
x = x+3
x += 3 # augmented operator
print(x)

x = 10 + 3 * 2 ** 2
print(x)
# Parenthesis
# Exponentioation
# Multiplication or division
# Addition or subtraction

x = (2+3) * 10 - 3
print(x)

x = 2.9
print(round(x))
print(abs(-2.9))

print(math.ceil(2.9))
print(math.floor(2.9))



is_hot = False
is_cold = False
if is_hot:
    print("It's a hot day")
    print("Drink plenty of water")
elif is_cold:
    print("It's a cold day, wear warm clothes")
else:
    print("It's a lovely day")

print("Enjoy your day")

price = 1000000
has_good_credit = True
down_payment = 0
if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f'Down payment is: [{down_payment}]')


# logical operators
has_high_income = True
has_good_credit = True
has_criminal_record = True

if has_good_credit and has_high_income and has_criminal_record:
    print("Eligible for loan")



# comparison operator
temp = 30

if temp > 30:
    print('Hot day')
elif temp < 5:
    print('Cold day')
else:
    print("it's a okay day")

name = "Anant Kishore"

if len(name) < 3:
    print('name is too short')
elif len(name) > 50:
    print('too big')
else:
    print(name)


# initialize and use it in while
x = 1
while x < 10:
    print('*'*x)
    x += 1

secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
    choice = int(input('Enter guess: '))
    guess_count += 1
    if choice == secret_number:
        print('you got it')
        break
else:
    print("no, bad luck")

command = ""
started = False
while True:
    command = input("> ").lower()
    if command == "start":
        if started:
            print("Car is already started")
        else:
            print("Car started...")
            started = True
    elif command == "stop":
        if not started:
            print("Car is already stopped")
        else:
            print("Car Stopped")
            started = False
    elif command == "help":
        print('''
        start - to start the car
        stop  - to stop the car
        quit  - to quit the program''')
    elif command == "quit":
        break
    else:
        print("it's invalid input")


for item in ['Anant', 'Kishore', 'Sarah']:
    print(item)

for item in range(10):
    print(item)

for item in [1,2,3,4]:
    print(item)

for item in range(5,10,2):
    print(item)

total = 0
prices = [10,20,30]
for price in prices:
    total += price;
print(f"Total: {total}")



# Nested Loops (x,y)

for x in range(4):
    for y in range(3):
        print(f'({x},{y})')

numbers = [5, 2, 5, 2, 2]

for item in numbers:
    print("#" * item)


names = ['Ananth', 'Kishore', 'Pankaj', 'Tiwari']
names[0] = 'Anant'
print(names[2:4])

numbers = [1,2,3,4,59,7]

max_number = numbers[0]

for item in numbers:
    if item > max_number:
        max_number = item
print(max_number)




matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
matrix[0][1] = 10
print(matrix[0][1])

for row in matrix:
    for col in matrix:
        print(col)

numbers = [5, 2, 1, 7, 4]

numbers.append(20)
print(numbers)

numbers.insert(0, 0)
print(numbers)

numbers.remove(7)
print(numbers)

item = numbers.pop()
print(item)
print(numbers)

print(numbers.index(2))
#print(numbers.index(134))

print(134 in numbers)

print(numbers.sort())
print(numbers)

numbers2 = numbers.copy()
numbers.append(10)

print(numbers2)
print(numbers)

numbers.clear()
print(numbers)

numbers = [2, 2, 4, 6, 3, 4, 6, 1]
uniques = []

for number in numbers:
    if number not in uniques:
        uniques.append(number)

print(uniques)


numbers = (1, 2, 3, 1)
print(numbers.count(1))
print(numbers.index(1))

coordinates = (3, 4, 7)

x = coordinates[0]
y = coordinates[1]
z = coordinates[2]

print(x)
print(y)
print(z)

#unpacking
x, y, z = coordinates

print(x)
print(y)
print(z)


#Dictionaries

customer = {
    "name": "Ananth Kishore",
    "age": 29,
    "is_employed": True,

}

customer["name"] = "Anant kishore"
print(customer["name"])
# None
print(customer.get("birth_year", "29 Sep 1989"))

customer["birth_year"] = "29/09/1989"

print(customer.get("birth_year", "29 Sep 1989"))

digits_mapping = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four"

}
number = input("Phone: ")

for item in number:
    print(digits_mapping[item], sep =" ", end = " ")

digits_mapping = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four"

}

phone = input("Phone: ")
result = ""

for ch in phone:
    result += digits_mapping.get(ch, "Unknown") + " "
else:
    print(result)

emojis = {
    ":)": "😖",
    ":(": "😍"
}
message = input("> ")

words = message.split(' ')

output = ""
for word in words:
    output += emojis.get(word, word)+ " "
else:
    print(output)


def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}')
    print('Welcome abroad')


print("Start")
greet_user("Anant", "Kishore")
print("Finish")

greet_user(last_name="Kishore", first_name="Anant")
greet_user("Kishore", last_name="Anant")


def sqaure(number):
    print(number * number)
#by default all functions return None

print(sqaure(3))

emojis = {
    ":)": "😖",
    ":(": "😍"
}
def get_mapping(ch):
    return emojis.get(ch, ch)

message = input("> ")

for ch in message.split(' '):
    print(get_mapping(ch), end=" ")

try:
    age = int(input('Age: '))
    income = 20000
    risk = income/age
    print(age)
except ValueError:
    print('Invalid value')
except ZeroDivisionError:
    print("Age can not be zero")

class Point:
    
    def __init__(self, x, y):
        self.x = x
        self.y = y
        
    def move(self):
        print("move")

    def draw(self):
        print("draw")


pt1 = Point()
pt1.draw()
pt1.move()
pt1.x = 10
pt1.y = 20

print(pt1.x)

pt2 = Point

#print(pt2.x)

pt2.x = 10

print(pt2.x)

class Point:

    def __init__(self, x, y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")

pt = Point(10, 20)
print(pt.x)

class Person:

    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f'Hi, I am {self.name}')

person = Person("Anant Kishore")

print(person.name)
person.talk()

person = Person("Pankaj Tiwari")

print(person.name)
person.talk()

#inheritance

class Animal:
    def walk(self):
        print("Animal walks")


class Dog(Animal):
    #pass to pass without body 
    def walk(self):
        print("bark")


class Cat(Animal):
    def walk(self):
        print("meow")

dog1 = Dog()
cat1 = Cat()

dog1.walk()
cat1.walk()


import converters
from converters import kgs_to_lb

print(kgs_to_lb(40))
print(converters.lbs_to_kg(100))

import findMax
numbers = [10, 34, 6, 2]
print(findMax.find_max_number(numbers))

from ecommerce import shipping

shipping.calc_shipping()


import random

for i in range(3):
    print(random.randint(10,100))

members = ['Anant' , 'Kishore', 'Bikki']

leader = random.choice(members)
print(leader)

choice1 = random.randint(1,6)
choice2 = random.randint(1,6)
print(f'({choice1}, {choice2})')


class Dice:
    def roll(self):
        choice1 = random.randint(1, 6)
        choice2 = random.randint(1, 6)
        return choice1, choice2 # it will go as tuple

dice = Dice()
print(dice.roll())

#File system

from pathlib import Path

#absolute path
#relative path

path = Path("ecommerce")
print(path.exists())

path = Path("emails")
path.mkdir()

path.rmdir()

path = Path()
print(path.glob('*.*'))

for file in path.glob('*'):
    print(file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    print(cell)
    cell = sheet.cell(1,1)
    print(cell.value)

    print(sheet.max_row)

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('transactions2.xlsx') #not sure why it doesn't work

process_workbook('transactions.xlsx')

"""






